from django.contrib import admin
from django import forms
from django.shortcuts import render, redirect
from django.core.exceptions import ValidationError
from .models import Book, Chapter, Section, Shloka, AudioFile, Role,User
from django.http import HttpResponse,FileResponse
from import_export import resources
from import_export.admin import ExportMixin
from django.urls import reverse
from django.urls import path
import openpyxl
from django.shortcuts import redirect
from django.contrib.sessions.models import Session
from django.contrib.admin.sites import site
from django.contrib.auth.admin import UserAdmin
from django.contrib.auth import logout
from django.http import HttpResponseRedirect
import os
from django.conf import settings


admin.site.site_header = "e-Sharir Admin-Panel"
admin.site.site_title = "e-Sharir"
admin.site.index_title = "Welcome to e-Sharir Book"


class ShlokaResource(resources.ModelResource):
    class Meta:
        model = Shloka
        fields = ('id', 'chapter', 'section', 'shloka_number', 'shlok_text')

class ExcelUploadForm(forms.Form):
    excel_file = forms.FileField(label="Upload Excel File",
                                 widget=forms.ClearableFileInput(attrs={"id": "id_excel_file"}), required=True)

class ShlokaAdminForm(forms.ModelForm):
    class Meta:
        model = Shloka
        fields = ['chapter', 'section', 'shloka_number', 'shlok_text', 'audio']

    def clean(self):
        cleaned_data = super().clean()
        if not cleaned_data.get('chapter'):
            raise ValidationError({'chapter': 'Chapter is required.'})
        if not cleaned_data.get('shloka_number'):
            raise ValidationError({'shloka_number': 'Shloka number is required.'})
        if not cleaned_data.get('shlok_text'):
            raise ValidationError({'shlok_text': 'Shlok text is required.'})
        return cleaned_data

class ShlokaAdmin(admin.ModelAdmin):
    form = ShlokaAdminForm
    list_display = ('id', 'shloka_number', 'shlok_text', 'audio', 'chapter', 'created_at', 'modified_at')
    change_list_template = 'books/change_list.html'

    search_fields = ('shlok_text', 'audio', 'chapter__chapter_name')
    list_filter = ('chapter', 'created_at', 'modified_at')

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('upload-shlokas/', self.admin_site.admin_view(self.upload_shlokas), name='upload-shlokas'),
            path('example-format/', self.admin_site.admin_view(self.example_format), name='shloka-example-format'),
        ]
        return custom_urls + urls

    def upload_shlokas(self, request):
        if request.method == 'POST':
            form = ExcelUploadForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = form.cleaned_data['excel_file']
                try:
                    workbook = openpyxl.load_workbook(excel_file)
                    sheet = workbook.active

                    headers = [cell.value for cell in sheet[1]]
                    expected_headers = ['shloka_number', 'shlok_text', 'audio', 'chapter_id']
                    if headers != expected_headers:
                        self.message_user(request, f"Invalid file format. Expected headers: {', '.join(expected_headers)}", level='error')
                        return redirect('..')

                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        shloka_number, shlok_text, audio, chapter_id = row
                        try:
                            chapter = Chapter.objects.get(id=chapter_id)
                            Shloka.objects.create(
                                chapter=chapter,
                                shloka_number=shloka_number,
                                shlok_text=shlok_text,
                                audio=audio
                            )
                        except Chapter.DoesNotExist:
                            self.message_user(request, f"Chapter with ID {chapter_id} does not exist.", level='error')
                        except Exception as e:
                            self.message_user(request, f"Error processing shloka {shloka_number}: {str(e)}", level='error')

                    self.message_user(request, "Shlokas uploaded successfully.", level='success')
                    return redirect('..')
                except Exception as e:
                    self.message_user(request, f"Error reading Excel file: {str(e)}", level='error')
                    return redirect('..')
        else:
            form = ExcelUploadForm()

        return render(request, 'admin/upload_excel.html', {'form': form})

    def example_format(self, request):
    # Create an example Excel file
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['shloka_number', 'shlok_text', 'audio', 'chapter_id'])
        sheet.append([1, 'Example Shloka', 'Example Audio', 1])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="shloka_example.xlsx"'
        workbook.save(response)
        return response


    actions = ['example_format']

class BookAdminForm(forms.ModelForm):
    class Meta:
        model = Book
        fields = ['book_number', 'book_name', 'book_image', 'book_slider']

    def clean(self):
        cleaned_data = super().clean()
        if not cleaned_data.get('book_number'):
            raise ValidationError({'book_number': 'Book number is required.'})
        if not cleaned_data.get('book_name'):
            raise ValidationError({'book_name': 'Book name is required.'})
        return cleaned_data

class BookAdmin(admin.ModelAdmin):
    form = BookAdminForm
    list_display = ('id', 'book_number', 'book_name', 'book_image', 'created_at', 'modified_at', 'modified_by', 'book_slider')
    search_fields = ('book_name',)
    list_filter = ('created_at', 'modified_at', 'modified_by')

class ChapterAdminForm(forms.ModelForm):
    class Meta:
        model = Chapter
        fields = ['book', 'chapter_number', 'chapter_name', 'chapter_slider']

    def clean(self):
        cleaned_data = super().clean()
        if not cleaned_data.get('book'):
            raise ValidationError({'book': 'Book is required.'})
        if not cleaned_data.get('chapter_number'):
            raise ValidationError({'chapter_number': 'Chapter number is required.'})
        if not cleaned_data.get('chapter_name'):
            raise ValidationError({'chapter_name': 'Chapter name is required.'})
        return cleaned_data

class ChapterAdmin(admin.ModelAdmin):
    form = ChapterAdminForm
    list_display = ('id', 'book', 'chapter_number', 'chapter_name', 'created_at', 'modified_at', 'modified_by', 'chapter_slider')
    search_fields = ('chapter_name', 'book__book_name')
    list_filter = ('book', 'created_at', 'modified_at')

class SectionAdmin(admin.ModelAdmin):
    list_display = ('id', 'chapter', 'section_number', 'section_name', 'section_image', 'created_at', 'modified_at', 'modified_by')
    search_fields = ('section_name', 'chapter__chapter_name')
    list_filter = ('chapter', 'created_at', 'modified_at')

class AudioFileAdmin(admin.ModelAdmin):
    list_display = ('id', 'file', 'text_file')

class RoleAdmin(admin.ModelAdmin):
    list_display = ('id', 'role_name', 'created_at', 'modified_at', 'modified_by')
    search_fields = ('role_name',)
    list_filter = ('created_at', 'modified_at')


class UserAdmin(UserAdmin):
    list_display = ('username', 'email', 'role', 'is_staff', 'is_active', 'date_joined', 'last_login')
    search_fields = ('username', 'email')
    list_filter = ('is_staff', 'is_active', 'date_joined', 'last_login')

# Register models
admin.site.register(User,UserAdmin)
if not site.is_registered(Shloka):
    admin.site.register(Shloka, ShlokaAdmin)
admin.site.register(Book, BookAdmin)
admin.site.register(Chapter, ChapterAdmin)
admin.site.register(Section, SectionAdmin)
admin.site.register(AudioFile, AudioFileAdmin)
admin.site.register(Role, RoleAdmin)
